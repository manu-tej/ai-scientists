"""Format adapters: read/perturb the answer-critical data per file type.

Each adapter is bound to one file path and exposes a uniform surface used by
both ops (perturbations) and the validator (reads). Unsupported write formats
raise UnsupportedFormat so the builder can flag rather than silently mis-build.
"""
from __future__ import annotations

import csv as _csv
from pathlib import Path

import pandas as pd


class UnsupportedFormat(Exception):
    pass


# --------------------------------------------------------------------------- #
# Tabular: csv / tsv / txt / xls / xlsx                                        #
# --------------------------------------------------------------------------- #

class TabularAdapter:
    DELIM = {".csv": ",", ".tsv": "\t", ".txt": None, ".diff": "\t"}  # None => sniff; .diff = Cuffdiff TSV

    def __init__(self, path: Path):
        self.path = Path(path)
        # Transparent gzip: a ".tsv.gz" file is a TSV that happens to be
        # compressed. Key the format/delimiter logic off the INNER extension
        # (.tsv) and route I/O through gzip, so ops/validator work unchanged.
        self.gz = self.path.suffix.lower() == ".gz"
        inner = Path(self.path.stem) if self.gz else self.path
        self.ext = inner.suffix.lower()
        self.is_excel = self.ext in (".xls", ".xlsx")

    def _open(self, mode: str):
        if self.gz:
            import gzip
            return gzip.open(self.path, mode + "t", newline="")
        return open(self.path, mode, newline="")

    # ---- delimiter sniff for plain text ---- #
    def _sep(self) -> str:
        s = self.DELIM.get(self.ext, ",")
        if s is not None:
            return s
        with self._open("r") as fh:
            first = fh.readline()
        return "\t" if ("\t" in first and "," not in first) else ","

    # ---- csv via the stdlib reader (robust to ragged preamble rows) ---- #
    def _rows(self) -> list[list[str]]:
        with self._open("r") as fh:
            return list(_csv.reader(fh, delimiter=self._sep()))

    def _write_rows(self, rows) -> None:
        # Resolve the delimiter BEFORE opening in "w" mode: open(...,"w")
        # truncates the file, after which _sep()'s first-line sniff would see
        # an empty line and wrongly fall back to ",". Snapshot it first so a
        # sniffed TSV (.txt) round-trips as TSV instead of being rewritten CSV.
        sep = self._sep()
        with self._open("w") as fh:
            _csv.writer(fh, delimiter=sep).writerows(rows)

    # ---- reads ---- #
    def list_columns(self, sheet=None, header_row=0) -> list:
        if self.is_excel:
            return list(pd.read_excel(self.path, sheet_name=sheet, header=header_row, nrows=1).columns)
        rows = self._rows()
        return [c for c in rows[header_row] if c != ""] if len(rows) > header_row else []

    def column_values(self, column, sheet=None, header_row=0) -> list:
        if self.is_excel:
            df = pd.read_excel(self.path, sheet_name=sheet, header=header_row)
            return df[column].dropna().astype(str).tolist() if column in df.columns else []
        rows = self._rows()
        header = rows[header_row]
        if column not in header:
            return []
        i = header.index(column)
        return [r[i] for r in rows[header_row + 1:] if len(r) > i and r[i] != ""]

    def n_rows(self, sheet=None, header_row=0) -> int:
        if self.is_excel:
            return len(pd.read_excel(self.path, sheet_name=sheet, header=header_row))
        return max(0, len(self._rows()) - header_row - 1)

    def sheet_names(self) -> list:
        return pd.ExcelFile(self.path).sheet_names if self.is_excel else ["<csv>"]

    # ---- writes ---- #
    def drop_columns(self, columns, sheet=None, header_row=0) -> None:
        targets = set(columns)
        if not self.is_excel:
            rows = self._rows()
            header = rows[header_row]
            drop_pos = {i for i, h in enumerate(header) if h in targets}
            self._write_rows([[v for i, v in enumerate(r) if i not in drop_pos] for r in rows])
            return
        if self.ext == ".xls":
            raise UnsupportedFormat(".xls is read-only; cannot drop columns (use the existing variant file)")
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        ws = wb[sheet]
        hdr = header_row + 1  # openpyxl is 1-based
        idxs = [cell.column for cell in ws[hdr] if cell.value in targets]
        for idx in sorted(idxs, reverse=True):
            ws.delete_cols(idx, 1)
        wb.save(self.path)

    def drop_columns_matching(self, pattern, sheet=None, header_row=0) -> None:
        import re
        rx = re.compile(pattern)
        cols = [c for c in self.list_columns(sheet=sheet, header_row=header_row) if rx.search(str(c))]
        if cols:
            self.drop_columns(cols, sheet=sheet, header_row=header_row)

    def keep_rows_where(self, column, keep_values, sheet=None, header_row=0) -> None:
        keep = {str(v) for v in keep_values}
        if not self.is_excel:
            rows = self._rows()
            header = rows[header_row]
            i = header.index(column)
            kept = [r for r in rows[header_row + 1:] if len(r) > i and r[i] in keep]
            self._write_rows(rows[:header_row + 1] + kept)
            return
        if self.ext == ".xls":
            raise UnsupportedFormat(".xls is read-only; cannot filter rows")
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        ws = wb[sheet]
        hdr = header_row + 1
        gcol = next((cell.column for cell in ws[hdr] if cell.value == column), None)
        if gcol is None:
            raise KeyError(f"column {column!r} not found in sheet {sheet!r}")
        for r in range(ws.max_row, hdr, -1):
            if str(ws.cell(row=r, column=gcol).value) not in keep:
                ws.delete_rows(r, 1)
        wb.save(self.path)

    def reduce_rows(self, n, seed=0, sheet=None, header_row=0) -> None:
        if self.is_excel and self.ext == ".xls":
            raise UnsupportedFormat(".xls is read-only; cannot subsample rows")
        if self.is_excel:
            df = pd.read_excel(self.path, sheet_name=sheet, header=header_row)
            df.sample(n=min(n, len(df)), random_state=seed).to_excel(self.path, sheet_name=sheet, index=False)
            return
        import random
        rows = self._rows()
        data = rows[header_row + 1:]
        rnd = random.Random(seed)
        keep = rnd.sample(data, k=min(n, len(data)))
        self._write_rows(rows[:header_row + 1] + keep)


# --------------------------------------------------------------------------- #
# AnnData: h5ad (obs-only edits via h5py; fast, no matrix rewrite)            #
# --------------------------------------------------------------------------- #

class AnnDataAdapter:
    def __init__(self, path: Path):
        self.path = Path(path)

    def _obs_colorder(self, obs) -> list:
        order = obs.attrs.get("column-order")
        if order is not None:
            return [c.decode() if isinstance(c, bytes) else str(c) for c in list(order)]
        return [k for k in obs.keys()]

    def obs_columns(self) -> list:
        import h5py
        with h5py.File(self.path, "r") as f:
            return self._obs_colorder(f["obs"])

    def obs_category_values(self, column) -> list:
        """Return the small set of distinct labels for a column (categories for
        categoricals; unique of the dataset otherwise)."""
        import h5py
        import numpy as np
        with h5py.File(self.path, "r") as f:
            if column not in f["obs"]:
                return []
            node = f["obs"][column]
            if isinstance(node, h5py.Group) and "categories" in node:  # categorical
                cats = node["categories"][:]
            else:
                cats = np.unique(node[:])
        return [c.decode() if isinstance(c, bytes) else str(c) for c in cats]

    def drop_obs_columns(self, columns) -> None:
        import h5py
        with h5py.File(self.path, "a") as f:
            obs = f["obs"]
            order = self._obs_colorder(obs)
            for col in columns:
                if col in obs:
                    del obs[col]
            new_order = [c for c in order if c not in set(columns)]
            obs.attrs["column-order"] = new_order

    def drop_obs_columns_matching(self, pattern) -> None:
        import re
        rx = re.compile(pattern)
        self.drop_obs_columns([c for c in self.obs_columns() if rx.search(str(c))])

    def anonymize_obs_column(self, column, prefix="id_") -> None:
        """Replace a column's labels with neutral IDs, preserving grouping.

        For a categorical, only the (small) categories array is rewritten, so the
        per-cell codes still define the same groups but the labels leak nothing.
        """
        import h5py
        with h5py.File(self.path, "a") as f:
            obs = f["obs"]
            if column not in obs:
                return
            node = obs[column]
            if isinstance(node, h5py.Group) and "categories" in node:
                n = node["categories"].shape[0]
                neutral = [f"{prefix}{i:04d}".encode() for i in range(n)]
                del node["categories"]
                node.create_dataset("categories", data=neutral)
            else:  # plain string dataset: map each distinct value to a neutral id
                import numpy as np
                vals = node[:]
                uniq = {v: f"{prefix}{i:04d}".encode() for i, v in enumerate(np.unique(vals))}
                mapped = np.array([uniq[v] for v in vals])
                del obs[column]
                obs.create_dataset(column, data=mapped)


def get_adapter(path: str | Path):
    p = Path(path)
    ext = p.suffix.lower()
    # Transparent gzip: dispatch on the inner extension (.tsv.gz -> .tsv).
    inner = Path(p.stem).suffix.lower() if ext == ".gz" else ext
    if inner in (".csv", ".tsv", ".txt", ".diff"):  # gz only for delimited text
        return TabularAdapter(p)
    if ext in (".csv", ".tsv", ".txt", ".diff", ".xls", ".xlsx"):
        return TabularAdapter(p)
    if ext == ".h5ad":
        return AnnDataAdapter(p)
    raise UnsupportedFormat(f"no adapter for '{ext}' ({p.name})")
